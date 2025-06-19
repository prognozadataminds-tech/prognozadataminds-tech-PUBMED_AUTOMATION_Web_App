import os
import time
import glob
import shutil
import pandas as pd
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait, Select
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.options import Options
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter

def process_excel(input_excel_path, output_excel_path):
    # Load Excel
    try:
        df = pd.read_excel(input_excel_path)
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        return

    if 'Comment' not in df.columns:
        df['Comment'] = ''

    if 'Result Count' not in df.columns:
        df['Result Count'] = ''

    # Setup Chrome Download Directory
    download_dir = os.path.join(os.getcwd(), "downloads")
    os.makedirs(download_dir, exist_ok=True)

    chrome_options = Options()
    chrome_options.add_experimental_option("prefs", {
        "download.default_directory": download_dir,
        "download.prompt_for_download": False,
        "download.directory_upgrade": True,
        "safebrowsing.enabled": True
    })
    chrome_options.add_argument("--start-maximized")

    # Launch Chrome
    driver = webdriver.Chrome(options=chrome_options)
    wait = WebDriverWait(driver, 20)

    try:
        for index, row in df.iterrows():
            search_query = row.get('Search Strategy 1')
            target_filename = row.get('File Name')

            if pd.isna(search_query) or pd.isna(target_filename):
                print(f"⏭️ Skipping row {index + 2} due to empty values")
                df.at[index, 'Comment'] = 'Empty values'
                continue

            print(f"\n🔍 Searching for: {search_query}")
            driver.get("https://pubmed.ncbi.nlm.nih.gov/")

            try:
                search_box = wait.until(EC.presence_of_element_located((By.ID, "id_term")))
                search_box.clear()
                search_box.send_keys(search_query)
                search_box.send_keys(Keys.ENTER)

                # Check if it's a search results page
                results_amount_elements = driver.find_elements(By.CLASS_NAME, "results-amount")
                is_search_results_page = len(results_amount_elements) > 0

                if is_search_results_page:
                    # Existing logic for search results page
                    wait.until(EC.presence_of_element_located((By.CLASS_NAME, "results-amount"))) # Ensure element is fully loaded
                    result_text = results_amount_elements[0].text
                    result_count = int(result_text.split()[0].replace(",", ""))
                    df.at[index, 'Result Count'] = result_count

                    if result_count > 1000:
                        print(f"⚠️ Too many results ({result_count}) — skipping download")
                        df.at[index, 'Comment'] = '1000+ Results - Not Downloaded'
                        continue
                    elif result_count == 0:
                        print(f"⚠️ No results found — skipping download")
                        df.at[index, 'Comment'] = 'No results found'
                        continue
                    else:
                        print(f"✅ Found {result_count} results — proceeding with download")

                    # Quoted warning
                    warnings = driver.find_elements(By.CLASS_NAME, "usa-alert-body")
                    for w in warnings:
                        if "Quoted phrase not found in" in w.text:
                            print("⚠️ Quoted phrase warning — skipping")
                            df.at[index, 'Comment'] = 'Quoted phrase warning'
                            raise Exception("Quoted phrase issue")

                else:
                    # Check if it's a single article page (e.g., by looking for "full-text-links")
                    full_text_links_elements = driver.find_elements(By.CLASS_NAME, "full-text-links")
                    if len(full_text_links_elements) > 0:
                        print("✅ Landed on single article page, proceeding with download.")
                        result_count = 1  # Set result_count for single article
                        df.at[index, 'Result Count'] = result_count
                    else:
                        # Neither search results nor single article page recognized
                        df.at[index, 'Comment'] = 'Search error or unexpected page layout'
                        print(f"⚠️ Search error or unexpected page layout after search for: {search_query}")
                        continue # Skip to next row

            except Exception as e:
                if not df.at[index, 'Comment']:
                    df.at[index, 'Comment'] = 'No results found'
                print(f"⚠️ Search error during initial page load/check: {e}")
                continue

            try:
                print("\n=== Starting save process ===")
                print(f"Current URL: {driver.current_url}")
                print(f"Result count: {result_count}")  # Debug print
                
                # Wait for and click save button
                print("Looking for save button...")
                save_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Save')]")))
                print("Found save button, clicking...")
                save_btn.click()
                print("Clicked save button")
                time.sleep(3)  # Increased wait time for dialog
                
                # Check if we're on the right page
                print("Checking page state...")
                if "save" not in driver.current_url.lower():
                    print("Warning: Not on save page after clicking save button")
                
                # Special handling for single result
                if result_count == 1:
                    print(f"\nProcessing single result (count: {result_count})")
                    try:
                        print("\n=== Starting format dropdown interaction ===")
                        print("Current URL:", driver.current_url)
                        
                        # Wait for save dialog with increased timeout
                        print("Waiting for save dialog...")
                        try:
                            save_dialog = WebDriverWait(driver, 30).until(
                                EC.presence_of_element_located((By.CLASS_NAME, "save-dialog"))
                            )
                            print("Save dialog found")
                        except Exception as dialog_error:
                            print(f"Save dialog not found: {str(dialog_error)}")
                            print("Trying alternative dialog detection...")
                            try:
                                save_dialog = WebDriverWait(driver, 30).until(
                                    EC.presence_of_element_located((By.CLASS_NAME, "usa-modal"))
                                )
                                print("Alternative save dialog found")
                            except Exception as alt_dialog_error:
                                print(f"Alternative dialog not found: {str(alt_dialog_error)}")
                                raise Exception("Could not find save dialog")

                        # Try multiple approaches to select PubMed format
                        format_selected = False
                        approaches = [
                            # Approach 1: Direct ID
                            lambda: driver.find_element(By.ID, "save-action-format").click(),
                            # Approach 2: Select class
                            lambda: Select(driver.find_element(By.ID, "save-action-format")).select_by_visible_text("PubMed"),
                            # Approach 3: Keyboard navigation
                            lambda: driver.find_element(By.ID, "save-action-format").send_keys(Keys.DOWN + Keys.RETURN)
                        ]

                        for i, approach in enumerate(approaches, 1):
                            try:
                                print(f"\nTrying approach {i}...")
                                approach()
                                time.sleep(2)  # Wait for selection to take effect
                                format_selected = True
                                print(f"Approach {i} succeeded")
                                break
                            except Exception as e:
                                print(f"Approach {i} failed: {str(e)}")
                                continue

                        if not format_selected:
                            raise Exception("All format selection approaches failed")

                        print("Format selection completed")
                        time.sleep(0.5)  # Increased wait time after format selection

                    except Exception as e:
                        print(f"\n⚠️ Error selecting citation format for single result: {str(e)}")
                        print("Current page source:")
                        print(driver.page_source[:500] + "...")  # Print first 500 chars of page source
                        df.at[index, 'Comment'] = 'Format selection error (single result)'
                        continue
                else:
                    print(f"\nProcessing multiple results (count: {result_count})")
                    # Original behavior for multiple results
                    try:
                        print("Waiting for save-action-selection...")
                        wait.until(EC.visibility_of_element_located((By.ID, "save-action-selection")))
                        print("Found save-action-selection")
                        driver.find_element(By.ID, "save-action-selection").send_keys(Keys.DOWN + Keys.RETURN)
                        print("Selected save-action-selection")
                        time.sleep(1)
                        print("Selecting format...")
                        driver.find_element(By.ID, "save-action-format").send_keys(Keys.DOWN + Keys.RETURN)
                        print("Selected format")
                        time.sleep(1)
                    except Exception as e:
                        print(f"Error in multiple results handling: {str(e)}")
                        raise

                # After handling format selection, click the Create file button
                print("\nLooking for Create file button...")
                create_btn = wait.until(EC.element_to_be_clickable((By.XPATH, "//button[contains(text(), 'Create file')]")))
                print("Found Create file button, clicking...")
                create_btn.click()
                print("Clicked Create file button")

                print("⏳ Waiting for file to download...")
                # Increased waiting time for download and added a loop for robustness
                download_complete = False
                for _ in range(5): # Check up to 5 times (total 15 seconds)
                    time.sleep(3)
                    list_of_files = glob.glob(os.path.join(download_dir, "pubmed-*.*"))
                    if list_of_files:
                        download_complete = True
                        break
                
                if download_complete:
                    latest_file = max(list_of_files, key=os.path.getctime)
                    new_path = os.path.join(download_dir, f"{target_filename}.txt")
                    shutil.move(latest_file, new_path)
                    print(f"✅ Downloaded: {target_filename}.txt")
                    df.at[index, 'Comment'] = 'Downloaded'
                else:
                    print("❌ No downloaded file found after waiting.")
                    df.at[index, 'Comment'] = 'Download error - file not found'

            except Exception as e:
                print(f"⚠️ Download failed: {e}")
                df.at[index, 'Comment'] = 'Download error'
                continue

    finally:
        driver.quit()
        print("\n🛑 Browser closed. Saving Excel...")

        df.to_excel(output_excel_path, index=False)

        try:
            wb = load_workbook(output_excel_path)
            ws = wb.active

            # Header Styling
            header_font = Font(bold=True, color="FFFFFF", size=12)
            header_fill = PatternFill(start_color="4F81BD", end_color="4F81BD", fill_type="solid")
            alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

            for col_num, col_cells in enumerate(ws.iter_cols(min_row=1, max_row=1), 1):
                for cell in col_cells:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = alignment
                    width = max(len(str(cell.value)) + 5, 15)
                    ws.column_dimensions[get_column_letter(col_num)].width = width

            # Data Cell Styling
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    cell.alignment = Alignment(wrap_text=True, vertical="top")

            wb.save(output_excel_path)
            print(f"✅ Excel styled and saved at: {output_excel_path}")

        except Exception as e:
            print(f"⚠️ Error styling Excel: {e}")

# ----------------------------
# ✅ Run the Function
# ----------------------------
if __name__ == "__main__":
    input_excel_path = "Pubs.xlsx"           # Ensure this file exists
    output_excel_path = "Pubs_Updated.xlsx"
    process_excel(input_excel_path, output_excel_path)
